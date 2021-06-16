from flask import Flask, render_template, request, redirect, url_for, flash
from flask_mysqldb import MySQL
from openpyxl import load_workbook
from flask_bootstrap import Bootstrap
from werkzeug.utils import secure_filename
import os
from os import path
from flask_wtf import CsrfProtect
UPLOAD_FOLDER = os.path.abspath("./tmp/")

app = Flask(__name__)
# Mysql Connection
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'bd_cargadatos'
csrf = CsrfProtect(app)
Bootstrap(app)
mysql = MySQL(app)
    
# Settings
app.secret_key= 'mysecretkey'
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

ALLOWED_EXTENSIONS = set(["xlsx","xls","mxltx","xltm"])


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1] in ALLOWED_EXTENSIONS

@app.route('/')
def Index():    
    data = obtenerDatos()
    return render_template('index.html', listado = data)

@app.errorhandler(404)
def page_not_found(error):
 return render_template("404.html"), 404

@app.route('/cargar', methods=['POST'])
def Cargar():
    if request.method == 'POST':   
        #carga de archivo
        nombre_archivo = request.files['archivo']        
        # nombre_archivo = request.form['archivo']
        # filename = secure_filename(nombre_archivo.filename)
        filename = nombre_archivo.filename
        nombre_archivo.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))

        #Se valida el formato de archivo permitido
        if nombre_archivo.filename and allowed_file(filename):

            wb = load_workbook(app.config["UPLOAD_FOLDER"]+'\\'+nombre_archivo.filename, read_only=True)

            # cursor = mysql.connection.cursor()                    
            #Hojas del libro
            sheets = wb.sheetnames
            for sheet in sheets:
                #Hoja activa
                sheetActive = wb[sheet]
                if sheet in wb.sheetnames and wb[sheet].max_column == 5 and wb[sheet].max_row > 0:                
                    #Obtener la cantidad de Columnas
                    cols = wb[sheet].max_column
                    #Obtener la cantidad de Filas
                    rows = wb[sheet].max_row
                    #Recorrer cada fila y obtener los datos de cada column                    
                    for nombre, apellido, nacionalidad, fechaContrato, sexo in sheetActive.iter_rows(min_row=2):
                        insertarDatos(nombre.value, apellido.value, nacionalidad.value, fechaContrato.value, sexo.value)                    
                    flash('La carga del archivo se ha ejecutado correctamente')

                else:
                    flash('Error al intentar cargar el archivo, está mal formado o no contiene información')

            return redirect(url_for('Index'))
        else:
            flash('Error al intentar cargar el archivo, extensión no permitida')
        return redirect(url_for('Index'))

def insertarDatos(nombre, apellido, nacionalidad, fechaContrato, sexo):
    cursor = mysql.connection.cursor()     
    query = '''INSERT INTO datos(nombre, apellido, nacionalidad, fechaContrato, sexo) values(%s,%s,%s,%s,%s)'''
    values = (nombre, apellido, nacionalidad, fechaContrato, sexo)                   
    cursor.execute(query, values)
    cursor.close()
    mysql.connection.commit()

def obtenerDatos():
    cursor =  mysql.connection.cursor()
    cursor.execute('''SELECT * FROM datos''')
    data = cursor.fetchall()
    return data

if __name__ == '__main__':
    app.run(port=3000, debug=True)